import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from repositories.device_repository import DeviceRepository
from repositories.charger_repository import ChargerRepository
from repositories.cable_repository import CableRepository

device_repo = DeviceRepository()
charger_repo = ChargerRepository()
cable_repo = CableRepository()

def build_excel():
    devices = device_repo.get_all_for_export()
    chargers = charger_repo.get_all_for_export()
    cables = cable_repo.get_all_for_export()
    d_stats = device_repo.get_stats()
    c_stats = charger_repo.get_stats()
    l_stats = cable_repo.get_stats()

    overview = {
        'total_phones': d_stats['phones'],
        'total_tablets': d_stats['tablets'],
        'total_chargers': c_stats['total'],
        'total_cables': l_stats['total'],
        'brands': d_stats['brands'],
        'by_model': d_stats['models'],
        'connectors': d_stats['connectors'],
        'charger_types': c_stats['charger_types'],
        'cable_types': l_stats['cable_types'],
    }

    wb = openpyxl.Workbook()

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    sub_fill = PatternFill("solid", fgColor="2E75B6")
    sub_font = Font(color="FFFFFF", bold=True)
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    good_fill = PatternFill("solid", fgColor="C6EFCE")
    broken_fill = PatternFill("solid", fgColor="FFC7CE")
    title_font = Font(size=16, bold=True, color="1F3864")
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def write_header(ws, row, cols, col_start=1):
        for j, col in enumerate(cols, col_start):
            c = ws.cell(row=row, column=j, value=col)
            c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = thin

    def write_row(ws, row_num, values, alt=False, fill=None):
        for j, val in enumerate(values, 1):
            c = ws.cell(row=row_num, column=j, value=val)
            c.border = thin
            c.alignment = Alignment(vertical='center', wrap_text=True)
            if fill:
                c.fill = fill
            elif alt:
                c.fill = alt_fill

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25

    ws.merge_cells('A1:D1')
    title = ws['A1']
    title.value = f"Stock Count Report — Generated {datetime.now().strftime('%d %b %Y %H:%M')}"
    title.font = title_font; title.alignment = center
    ws.row_dimensions[1].height = 35

    row = 3
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "OVERALL TOTALS"
    ws[f'A{row}'].fill = sub_fill; ws[f'A{row}'].font = sub_font; ws[f'A{row}'].alignment = center
    row += 1
    for label, val in [('Total Phones', overview['total_phones']),
                        ('Total Tablets', overview['total_tablets']),
                        ('Total Chargers', overview['total_chargers']),
                        ('Total Cables', overview['total_cables']),
                        ('Total Devices', overview['total_phones'] + overview['total_tablets'])]:
        ws.cell(row=row, column=1, value=label).border = thin
        c = ws.cell(row=row, column=2, value=val)
        c.border = thin; c.font = Font(bold=True, size=12)
        row += 1

    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "BY BRAND"
    ws[f'A{row}'].fill = sub_fill; ws[f'A{row}'].font = sub_font; ws[f'A{row}'].alignment = center
    row += 1
    write_header(ws, row, ['Brand', 'Count'])
    row += 1
    for i, b in enumerate(overview['brands']):
        write_row(ws, row, [b['brand'], b['c']], alt=(i % 2 == 1))
        row += 1

    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "BY MODEL"
    ws[f'A{row}'].fill = sub_fill; ws[f'A{row}'].font = sub_font; ws[f'A{row}'].alignment = center
    row += 1
    write_header(ws, row, ['Brand', 'Model', 'Type', 'Count'])
    row += 1
    for i, m in enumerate(overview['by_model']):
        write_row(ws, row, [m['brand'], m['model'], m['device_type'], m['c']], alt=(i % 2 == 1))
        row += 1

    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "CONNECTOR ANALYSIS (Devices)"
    ws[f'A{row}'].fill = sub_fill; ws[f'A{row}'].font = sub_font; ws[f'A{row}'].alignment = center
    row += 1
    write_header(ws, row, ['Connector Type', 'Devices'])
    row += 1
    conn_map = {r['connector']: r['c'] for r in overview['connectors']}
    for conn, cnt in conn_map.items():
        write_row(ws, row, [conn, cnt])
        row += 1

    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "CABLE ↔ CHARGER MATCHING ANALYSIS"
    ws[f'A{row}'].fill = sub_fill; ws[f'A{row}'].font = sub_font; ws[f'A{row}'].alignment = center
    row += 1
    write_header(ws, row, ['Category', 'Devices / Cables Needed', 'Have', 'Gap (Need to Order)'])
    row += 1
    ct_map = {r['cable_type']: r['c'] for r in overview['cable_types']}
    chr_map = {r['charger_type']: r['c'] for r in overview['charger_types']}
    lightning_d = conn_map.get('lightning', 0)
    usbc_d = conn_map.get('USB-C', 0)
    microusb_d = conn_map.get('micro-USB', 0)
    lightning_c = ct_map.get('USB-C to Lightning', 0) + ct_map.get('USB-A to Lightning', 0)
    usbc_c = ct_map.get('USB-C to USB-C', 0) + ct_map.get('USB-A to USB-C', 0)
    microusb_c = ct_map.get('USB-A to micro-USB', 0)
    usba_chargers = chr_map.get('USB-A', 0) + chr_map.get('USB-C+USB-A', 0)
    usbc_chargers = chr_map.get('USB-C', 0) + chr_map.get('USB-C+USB-A', 0)
    usba_cables_count = sum(v for k, v in ct_map.items() if k.startswith('USB-A'))
    usbc_cables_charge = sum(v for k, v in ct_map.items() if k.startswith('USB-C to'))
    matching_rows = [
        ('Lightning Cables needed (for Lightning devices)', lightning_d, lightning_c, lightning_d - lightning_c),
        ('USB-C Cables needed (for USB-C devices)', usbc_d, usbc_c, usbc_d - usbc_c),
        ('micro-USB Cables needed', microusb_d, microusb_c, microusb_d - microusb_c),
        ('USB-A Chargers (for USB-A cables)', usba_cables_count, usba_chargers, usba_cables_count - usba_chargers),
        ('USB-C Chargers (for USB-C cables)', usbc_cables_charge, usbc_chargers, usbc_cables_charge - usbc_chargers),
    ]
    for label, needed, have, gap in matching_rows:
        fill = broken_fill if gap > 0 else good_fill
        for j, val in enumerate([label, needed, have, gap if gap > 0 else '✓ OK'], 1):
            c = ws.cell(row=row, column=j, value=val)
            c.fill = fill; c.border = thin; c.alignment = Alignment(vertical='center')
        row += 1

    # Sheet 2: Phones
    ws2 = wb.create_sheet("Phones")
    phones_data = [d for d in devices if d['device_type'] == 'phone']
    cols = ['Internal Barcode', 'Brand', 'Model', 'Connector', 'Engraved', 'Status', 'Place', 'Added', 'Added By', 'Updated', 'Updated By']
    ws2.merge_cells(f'A1:K1')
    ws2['A1'] = f"Phones ({len(phones_data)} total)"
    ws2['A1'].font = title_font; ws2['A1'].alignment = center
    write_header(ws2, 2, cols)
    for i, d in enumerate(phones_data, 3):
        write_row(ws2, i, [d['internal_barcode'], d['brand'], d['model'], d['connector'],
                           'Yes' if d['is_engraved'] else 'No', d['status'], d.get('place', ''),
                           d.get('created_at', ''), d.get('created_by_name', ''),
                           d.get('updated_at', ''), d.get('updated_by_name', '')],
                  fill=broken_fill if d['status'] == 'Broken' else (alt_fill if i % 2 == 0 else None))
    for j, w in enumerate([20, 15, 20, 12, 10, 10, 20, 20, 20, 20, 20], 1):
        ws2.column_dimensions[get_column_letter(j)].width = w
    ws2.auto_filter.ref = "A2:K2"

    # Sheet 3: Tablets
    ws3 = wb.create_sheet("Tablets")
    tablets_data = [d for d in devices if d['device_type'] == 'tablet']
    ws3.merge_cells(f'A1:K1')
    ws3['A1'] = f"Tablets ({len(tablets_data)} total)"
    ws3['A1'].font = title_font; ws3['A1'].alignment = center
    write_header(ws3, 2, cols)
    for i, d in enumerate(tablets_data, 3):
        write_row(ws3, i, [d['internal_barcode'], d['brand'], d['model'], d['connector'],
                           'Yes' if d['is_engraved'] else 'No', d['status'], d.get('place', ''),
                           d.get('created_at', ''), d.get('created_by_name', ''),
                           d.get('updated_at', ''), d.get('updated_by_name', '')],
                  fill=broken_fill if d['status'] == 'Broken' else (alt_fill if i % 2 == 0 else None))
    for j, w in enumerate([20, 15, 20, 12, 10, 10, 20, 20, 20, 20, 20], 1):
        ws3.column_dimensions[get_column_letter(j)].width = w
    ws3.auto_filter.ref = "A2:K2"

    # Sheet 4: Chargers
    ws4 = wb.create_sheet("Chargers")
    ws4.merge_cells('A1:G1')
    ws4['A1'] = f"Chargers ({len(chargers)} total)"
    ws4['A1'].font = title_font; ws4['A1'].alignment = center
    write_header(ws4, 2, ['ID', 'Barcode', 'Type', 'Place', 'Added', 'Added By', 'Updated By'])
    for i, c in enumerate(chargers, 3):
        write_row(ws4, i, [c['id'], c.get('barcode', ''), c['charger_type'], c.get('place', ''),
                           c.get('created_at', ''), c.get('created_by_name', ''), c.get('updated_by_name', '')],
                  alt=(i % 2 == 0))
    for j, w in enumerate([8, 20, 25, 20, 20, 20, 20], 1):
        ws4.column_dimensions[get_column_letter(j)].width = w
    ws4.auto_filter.ref = "A2:G2"

    # Sheet 5: Cables
    ws5 = wb.create_sheet("Cables")
    ws5.merge_cells('A1:G1')
    ws5['A1'] = f"Cables ({len(cables)} total)"
    ws5['A1'].font = title_font; ws5['A1'].alignment = center
    write_header(ws5, 2, ['ID', 'Barcode', 'Type', 'Place', 'Added', 'Added By', 'Updated By'])
    for i, c in enumerate(cables, 3):
        write_row(ws5, i, [c['id'], c.get('barcode', ''), c['cable_type'], c.get('place', ''),
                           c.get('created_at', ''), c.get('created_by_name', ''), c.get('updated_by_name', '')],
                  alt=(i % 2 == 0))
    for j, w in enumerate([8, 20, 28, 20, 20, 20, 20], 1):
        ws5.column_dimensions[get_column_letter(j)].width = w
    ws5.auto_filter.ref = "A2:G2"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out